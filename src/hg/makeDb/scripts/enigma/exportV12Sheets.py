#RM#38130
# Export the sheets needed for the v1.2 track rebuild from the CSpec registry xlsx
# files to tab-separated text. Merged cell ranges are expanded (top-left value copied
# to every cell in the range) because the v1.2 sheets use vertical merges for notes
# and warnings that span groups of rows.
#
# Inputs (downloaded from the CSpec registry Files & Images panel, see makedoc):
#   /hive/data/inside/enigmaTracksData/v1.2/CSpec_BRCA12ACMG_Rules-Specifications_V1.2_Table-9.xlsx
#   /hive/data/inside/enigmaTracksData/v1.2/CSpec_BRCA12ACMG_Rules-Specifications_V1.2_Table-4.xlsx
# Outputs:
#   .../v1.2/CSpec_BRCA12ACMG_Rules-Specifications_V1.2_Table-9.txt  (1.1-style layout,
#       consumed by BRCAfunctionalAssays.py: title row, header row, column-description
#       row, then data - the banner row 2 that v1.2 inserted is dropped)
#   .../v1.2/Table4_V1.2_annotatedExons.tsv  (consumed by convertTable4toFlat.py)
#
# The Table 9 xlsx also carries a "Dace & Findlay, Interim Report" sheet with interim
# (uncalibrated) assay results. It is intentionally not exported; the track shows the
# calibrated Table 9 only.

import openpyxl

workDir = "/hive/data/inside/enigmaTracksData/v1.2/"

def cleanCell(value):
    if value is None:
        return ""
    text = str(value).replace("\n", " ").replace("\t", " ").replace("▼", "")
    return text

def expandMerges(ws):
    ranges = list(ws.merged_cells.ranges)
    for mr in ranges:
        ws.unmerge_cells(str(mr))
        topLeft = ws.cell(row=mr.min_row, column=mr.min_col).value
        for row in range(mr.min_row, mr.max_row + 1):
            for col in range(mr.min_col, mr.max_col + 1):
                ws.cell(row=row, column=col).value = topLeft
    return len(ranges)

def sheetToRows(ws):
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([cleanCell(c) for c in row])
    return rows

# Table 9
wb = openpyxl.load_workbook(workDir + "CSpec_BRCA12ACMG_Rules-Specifications_V1.2_Table-9.xlsx", data_only=True)
ws = wb["Table9_BRCA12VCEP_specs"]
merged = expandMerges(ws)
rows = sheetToRows(ws)
# v1.2 layout: row1 title, row2 Dace & Findlay banner, row3 header, row4 column
# descriptions, row5+ data. Drop the banner so the output matches the 1.1 layout
# (title, header, descriptions, data) that BRCAfunctionalAssays.py skips with tail -n +4.
del rows[1]
with open(workDir + "CSpec_BRCA12ACMG_Rules-Specifications_V1.2_Table-9.txt", "w") as f:
    for row in rows:
        f.write("\t".join(row) + "\n")
print("Table 9: %d merged ranges expanded, %d rows written (banner row dropped)" % (merged, len(rows)))

# Table 4
wb = openpyxl.load_workbook(workDir + "CSpec_BRCA12ACMG_Rules-Specifications_V1.2_Table-4.xlsx", data_only=True)
ws = wb["Table 4 - Annotated Exons"]
merged = expandMerges(ws)
rows = sheetToRows(ws)
with open(workDir + "Table4_V1.2_annotatedExons.tsv", "w") as f:
    for row in rows:
        f.write("\t".join(row) + "\n")
print("Table 4: %d merged ranges expanded, %d rows written" % (merged, len(rows)))
