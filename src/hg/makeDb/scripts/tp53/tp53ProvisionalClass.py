#!/usr/bin/env python3
"""
TP53 VCEP NON-FINAL Provisional Classification track generator.

For every possible p53 missense protein change, sum the Tavtigian points
from applicable evidence codes the VCEP has pre-computed:

  * PM1 hotspot         (Clinical Domains + cancerhotspots.org)
  * PS3 / BS3           (CSpec GN009 v2.4.0 Table S3)
  * PP3 / BP4           (CSpec GN009 v2.4.0 Table S2)
  * BA1 / BS1 / PM2     (gnomAD v4.1 exomes, per CSpec FAF/grpmax thresholds)
  * BS2                 (FLOSSIES healthy-women-over-70 cohort)
  * Splicing PP3        (SpliceAI >= 0.2)

The point sum is bucketed into P / LP / VUS / LB / B per the CSpec
classification ranges. BA1 is stand-alone Benign and forces class = Benign
regardless of other evidence.

DELIBERATELY EXCLUDED from the sum (documented in every mouseover):
  - PVS1 (null variants only; handled in separate track)
  - PS1 / PS2 / PS4 / PP1 / PP4 / BS4 (require clinical observations)
  - BP7 (computational, but synonymous/intronic only; out of scope for this
    missense-only track)

This is NOT a ClinGen classification &#8212; the warning is in every mouseover
since clinicians live in the mouseover, not the description page.

bigBed 9+11.
"""

import argparse
import json
import os
import re
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import tp53FuncLib as lib

DEFAULT_OUTDIR = "/hive/users/lrnassar/claude/RM37399/provisionalClass"
SRC_S3 = "/hive/users/lrnassar/claude/RM37399/tp53_downloads/Functional-worksheet.xlsx"
SRC_S2 = "/hive/users/lrnassar/claude/RM37399/tp53_downloads/bioinformatic_worksheet.xlsx"
HOTSPOTS_JSON = "/hive/users/lrnassar/claude/RM37399/cancerHotspots/cancerhotspots_single.json"
AF_BED_TPL = "/hive/users/lrnassar/claude/RM37399/afFrequencies/TP53AF_{}.bed"
AF_PRESENT_TPL = "/hive/users/lrnassar/claude/RM37399/afFrequencies/TP53AF_present_{}.txt"
FLOSSIES_BED_TPL = "/hive/users/lrnassar/claude/RM37399/flossies/TP53Flossies_{}.bed"

PM1_HARDCODED_CODONS = {175, 245, 248, 249, 273, 282}

PROT_RE = re.compile(r'^([A-Z])(\d+)([A-Z])$')
HGVSP3_RE = re.compile(r'^p\.([A-Z][a-z]{2})(\d+)([A-Z][a-z]{2})$')
HGVSC_RE = re.compile(r'^c\.(\d+)([ACGT])>([ACGT])$')

THREE_TO_ONE = {
    'Ala':'A','Arg':'R','Asn':'N','Asp':'D','Cys':'C','Glu':'E','Gln':'Q',
    'Gly':'G','His':'H','Ile':'I','Leu':'L','Lys':'K','Met':'M','Phe':'F',
    'Pro':'P','Ser':'S','Thr':'T','Trp':'W','Tyr':'Y','Val':'V',
}

# Tavtigian points per code
POINTS = {
    'PM1_Moderate': 2, 'PM1_Supporting': 1,
    'PS3_Strong': 4, 'PS3': 4, 'PS3_Moderate': 2, 'PS3_Supporting': 1,
    'BS3_Strong': -4, 'BS3': -4, 'BS3_Supporting': -1,
    'PP3_Moderate': 2, 'PP3_moderate': 2, 'PP3': 1,
    'BP4_Moderate': -2, 'BP4_moderate': -2, 'BP4': -1,
    'BA1': 0,                   # stand-alone B -> not summed; forces class
    'BS1': -4,
    'PM2_Supporting': 1,
    'No evidence': 0, 'Indeterminate': 0,
}

SPLICE_PP3_THRESHOLD = 0.2  # CSpec GN009: SpliceAI >= 0.2 -> PP3 splicing


def bucket(pts):
    if pts >= 10:
        return 'Pathogenic'
    if 6 <= pts <= 9:
        return 'Likely pathogenic'
    if -6 <= pts <= -2:
        return 'Likely benign'
    if pts <= -7:
        return 'Benign'
    return 'Uncertain significance'


CLASS_COLOR = {
    'Pathogenic':              '210,0,0',
    'Likely pathogenic':       '245,152,152',
    'Uncertain significance':  '0,0,136',
    'Likely benign':           '213,247,213',
    'Benign':                  '0,210,0',
}

AUTOSQL = """table TP53ProvisionalClass
"TP53 NON-FINAL Provisional Classification by Tavtigian point sum (NOT a ClinGen classification)"
   (
   string chrom;             "Reference sequence chromosome or scaffold"
   uint   chromStart;        "Start position in chromosome"
   uint   chromEnd;          "End position in chromosome"
   string name;              "Missense change (e.g., R175H)"
   uint   score;             "Not used, all 0"
   char[1] strand;           "Not used, all ."
   uint   thickStart;        "Same as chromStart"
   uint   thickEnd;          "Same as chromEnd"
   uint   reserved;          "RGB color"
   string provisionalClass;  "Provisional class (Pathogenic/LP/VUS/LB/Benign)"
   int    totalPoints;       "Sum of applied Tavtigian points"
   string appliedCodes;      "Codes contributing (semicolon-separated)"
   string pm1;               "PM1 contribution"
   string ps3bs3;            "PS3/BS3 contribution (from Table S3)"
   string pp3bp4;            "PP3/BP4 contribution (Table S2 + splicing)"
   string af;                "AF code (BA1/BS1/PM2_Supporting/none)"
   string bs2;               "BS2 evidence (FLOSSIES cohort observation)"
   string spliceAI;          "Max SpliceAI delta (from Table S2)"
   string ntVerify;          "nt-divergence flag: SpliceAI / gnomAD AF path-dependent, verify per-nt"
   lstring _mouseOver;       "HTML mouseover"
   )
"""


def parse_missense_short(p):
    if not isinstance(p, str):
        return None
    m = PROT_RE.match(p.strip())
    if not m:
        return None
    return (m.group(1), int(m.group(2)), m.group(3))


def parse_hgvsp3(p):
    if not isinstance(p, str):
        return None
    m = HGVSP3_RE.match(p.strip())
    if not m:
        return None
    wt3, codon, alt3 = m.group(1), int(m.group(2)), m.group(3)
    if wt3 not in THREE_TO_ONE or alt3 not in THREE_TO_ONE:
        return None
    return (THREE_TO_ONE[wt3], codon, THREE_TO_ONE[alt3])


def load_s3(path):
    """Table S3: per-missense preliminary PS3/BS3. Key by (wt, codon, alt) 1-letter."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Supplementary Table S3"]
    out = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        k = parse_missense_short(row[0])
        if not k:
            continue
        out[k] = {
            'code': str(row[7]).strip() if row[7] is not None else 'No evidence',
        }
    return out


def load_s2(path):
    """Table S2: per c.X>Y missense with preliminary PP3/BP4 + max SpliceAI.

    Returns {(wt, codon, alt) -> {code, spliceai (max), spliceai_min,
    paths: [(c_pos, ref_nt, alt_nt), ...]}}.
    'paths' enumerates every c.X>Y combination that yields the protein change &#8212;
    used downstream to look up AF at each contributing genomic coord.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Supplementary Table S2"]
    out = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        k = parse_hgvsp3(row[1])
        if not k:
            continue
        hgvsc = row[0]
        path_tuple = None
        if isinstance(hgvsc, str):
            mc = HGVSC_RE.match(hgvsc.strip())
            if mc:
                path_tuple = (int(mc.group(1)), mc.group(2), mc.group(3))
        new_code = str(row[4]).strip() if row[4] is not None else 'No evidence'
        try:
            new_spliceai = float(row[5]) if row[5] is not None else 0.0
        except (ValueError, TypeError):
            new_spliceai = 0.0
        existing = out.get(k)
        if existing is None:
            out[k] = {
                'code': new_code,
                'spliceai': new_spliceai,      # max across nt paths
                'spliceai_min': new_spliceai,  # min across nt paths (straddle test)
                'paths': [path_tuple] if path_tuple else [],
            }
        else:
            if _s2_priority(new_code) > _s2_priority(existing['code']):
                existing['code'] = new_code
            if new_spliceai > existing['spliceai']:
                existing['spliceai'] = new_spliceai
            if new_spliceai < existing['spliceai_min']:
                existing['spliceai_min'] = new_spliceai
            if path_tuple and path_tuple not in existing['paths']:
                existing['paths'].append(path_tuple)
    return out


def _s2_priority(code):
    order = {
        'PP3_moderate': 5, 'BP4_moderate': 5,
        'PP3': 4,          'BP4': 4,
        'No evidence': 1,
    }
    return order.get(code, 0)


def load_hotspot_occurrences(path):
    out = {}
    if not os.path.exists(path):
        return out
    with open(path) as f:
        data = json.load(f)
    for h in data:
        if h.get('hugoSymbol') != 'TP53':
            continue
        res = h.get('residue') or ''
        if not res:
            continue
        wt = res[0]
        try:
            codon = int(res[1:])
        except ValueError:
            continue
        for alt, count in (h.get('variantAminoAcid') or {}).items():
            if alt == wt or alt in ('*', 'X') or count is None or count < 2:
                continue
            out[(codon, alt)] = count
    return out


def load_af_lookup(db):
    """Read TP53AF_<db>.bed and return {(chrom, pos1based, ref, alt) -> code}.

    The display name in the AF bed is 'chrN-pos-ref-alt' (1-based pos).
    """
    out = {}
    path = AF_BED_TPL.format(db)
    if not os.path.exists(path):
        return out
    with open(path) as f:
        for line in f:
            flds = line.rstrip("\n").split("\t")
            if len(flds) < 11:
                continue
            disp = flds[3]
            code = flds[9]
            m = re.match(r'^(chr[^-]+)-(\d+)-([ACGT-]+)-([ACGT-]+)$', disp)
            if not m:
                continue
            out[(m.group(1), int(m.group(2)), m.group(3), m.group(4))] = code
    return out


def load_gnomad_present(db):
    """Read TP53AF_present_<db>.txt (one 'chrN-pos-ref-alt' key per line, 1-based)
    and return a set of (chrom, pos1based, ref, alt) for every gnomAD variant at
    the TP53 locus. Used to tell 'absent from gnomAD' from 'present but not
    coded' when deciding PM2_Supporting."""
    out = set()
    path = AF_PRESENT_TPL.format(db)
    if not os.path.exists(path):
        return out
    with open(path) as f:
        for line in f:
            line = line.strip()
            m = re.match(r'^(chr[^-]+)-(\d+)-([ACGT-]+)-([ACGT-]+)$', line)
            if m:
                out.add((m.group(1), int(m.group(2)), m.group(3), m.group(4)))
    return out


BS2_TIER_POINTS = {'BS2': -4, 'BS2_Moderate': -2, 'BS2_Supporting': -1}


def load_flossies_lookup(db):
    """Read TP53Flossies_<db>.bed and return {(chrom, pos1based, ref, alt) ->
    (bs2_label, points)} for rows that meet BS2 at any tier (BS2 / BS2_Moderate
    / BS2_Supporting). Match is variant-specific (a FLOSSIES observation
    supports BS2 only for the exact nt change observed, not for any change at
    the same codon). The tier reflects the carrier count per CSpec GN009."""
    out = {}
    path = FLOSSIES_BED_TPL.format(db)
    if not os.path.exists(path):
        return out
    with open(path) as f:
        for line in f:
            flds = line.rstrip("\n").split("\t")
            if len(flds) < 19:
                continue
            pts = BS2_TIER_POINTS.get(flds[9])
            if pts is None:
                continue
            chrom = flds[0]
            start = int(flds[1])
            ref = flds[17]
            alt = flds[18]
            out[(chrom, start + 1, ref, alt)] = (flds[9], pts)
    return out


def pm1_code_and_points(wt, codon, alt, hotspot_occ):
    if codon in PM1_HARDCODED_CODONS:
        return ('PM1_Moderate', 2)
    count = hotspot_occ.get((codon, alt), 0)
    if count >= 10:
        return ('PM1_Moderate', 2)
    if count >= 2:
        return ('PM1_Supporting', 1)
    return ('', 0)


def ps3_bs3_label_and_points(code):
    if code in ('PS3', 'PS3_Strong'):
        return ('PS3_Strong', 4)
    if code == 'PS3_Moderate':
        return (code, 2)
    if code == 'PS3_Supporting':
        return (code, 1)
    if code == 'BS3_Supporting':
        return (code, -1)
    if code in ('BS3', 'BS3_Strong'):
        return ('BS3_Strong', -4)
    return ('', 0)


def pp3_bp4_label_and_points(code):
    if code == 'PP3_moderate':
        return ('PP3_Moderate', 2)
    if code == 'PP3':
        return ('PP3', 1)
    if code == 'BP4':
        return ('BP4', -1)
    if code == 'BP4_moderate':
        return ('BP4_Moderate', -2)
    return ('', 0)


def af_code_and_points(wt, codon, alt, paths, af_lookup, present_set, tx):
    """For a (wt, codon, alt) protein change, look up gnomAD v4.1 AF at every
    c.X>Y path and return the strongest applicable code.

    Priority (most-benign first): BA1 > BS1 > PM2_Supporting > none.
    BA1 is stand-alone Benign; the caller forces classification = Benign.

    PM2_Supporting applies either when a path is coded PM2 (present but rare) or
    when every path is absent from gnomAD (present_set). A path that is present
    in gnomAD but not rare enough to be coded blocks the absent-based PM2, so a
    variant seen in gnomAD at moderate frequency is not given PM2.

    Also returns af_divergent: True when the nt paths yield different AF codes
    (e.g. one path absent from gnomAD -> PM2 while another is present at moderate
    frequency -> no code). The protein-level row can only show one, so these are
    flagged for per-nt verification in the AF track.
    """
    chrom = tx['chrom']
    strand = tx['strand']
    rcomp = {'A':'T','T':'A','C':'G','G':'C'}
    per_path = []   # per nt path: 'BA1' / 'BS1' / 'PM2_Supporting' / 'absent' / 'other'
    for c_pos, c_ref, c_alt in paths:
        g = lib.cdna_coding_to_genomic(c_pos, tx)
        if g is None:
            continue
        if strand == '-':
            g_ref = rcomp.get(c_ref, c_ref)
            g_alt = rcomp.get(c_alt, c_alt)
        else:
            g_ref = c_ref
            g_alt = c_alt
        # AF lookup uses 1-based start
        key = (chrom, g + 1, g_ref, g_alt)
        code = af_lookup.get(key)
        if code:
            per_path.append(code)
        elif key not in present_set:
            per_path.append('absent')   # absent from gnomAD -> PM2-eligible
        else:
            per_path.append('other')    # present but not rare enough -> no code
    coded = [c for c in per_path if c not in ('absent', 'other')]
    any_absent = 'absent' in per_path
    any_present = ('other' in per_path) or bool(coded)

    # Effective code per path (absent -> PM2, present-Other -> none); if the set
    # disagrees the protein change is path-dependent for allele frequency.
    eff = set('PM2_Supporting' if s == 'absent' else ('' if s == 'other' else s)
              for s in per_path)
    af_divergent = len(eff) > 1

    if 'BA1' in coded:
        return ('BA1', 0, 'stand-alone Benign', af_divergent)
    if 'BS1' in coded:
        return ('BS1', POINTS['BS1'], '-4 pts', af_divergent)
    if 'PM2_Supporting' in coded:
        return ('PM2_Supporting', POINTS['PM2_Supporting'], '+1 pt', af_divergent)
    if any_absent and not any_present:
        return ('PM2_Supporting', POINTS['PM2_Supporting'],
                '+1 pt (absent from gnomAD)', af_divergent)
    return ('', 0, '', af_divergent)


def bs2_evidence(wt, codon, paths, flossies_lookup, tx):
    """Return (bs2_label, points) for the strongest FLOSSIES BS2 tier matching
    any (c.X>Y) path by exact genomic position AND ref/alt, else (None, 0).
    BS2 requires the SAME nt change &#8212; an observation of c.1120G>A does
    not support BS2 for c.1120G>C even though both yield p.G374R."""
    chrom = tx['chrom']
    strand = tx['strand']
    rcomp = {'A':'T','T':'A','C':'G','G':'C'}
    best = (None, 0)
    for c_pos, c_ref, c_alt in paths:
        g = lib.cdna_coding_to_genomic(c_pos, tx)
        if g is None:
            continue
        if strand == '-':
            g_ref = rcomp.get(c_ref, c_ref)
            g_alt = rcomp.get(c_alt, c_alt)
        else:
            g_ref = c_ref
            g_alt = c_alt
        hit = flossies_lookup.get((chrom, g + 1, g_ref, g_alt))
        if hit and hit[1] < best[1]:
            best = hit
    return best


CAVEATS_STR = (
    "PVS1 (null variants only &#8212; see PVS1 tracks); "
    "PS1/PS2/PS4/PP1/PP4/BS4 (require clinical observations); "
    "BP7 (synonymous/intronic; N/A to missense)."
)


HEADER_WARNING = (
    "<b style='color:#c00000'>"
    "NON-FINAL: NOT a ClinGen classification &#8212; preliminary point-sum only."
    "</b>"
)


def mouseover(name, cls, pts, pm1, ps3, pp3, af_lbl, bs2_lbl, applied,
              spliceai, splice_pp3_active, splice_overrode_bp4, codon, ba1,
              spliceai_min=0.0, splice_straddle=False, af_divergent=False,
              ps3_suppressed=''):
    splice_section = ""
    if splice_straddle:
        splice_section = (
            "<br><b>SpliceAI:</b> {mn:.2f}&ndash;{mx:.2f} &mdash; "
            "<span style='color:#c00000'>path-dependent: the nt variants at this "
            "codon straddle the 0.2 threshold, so splicing PP3 is not applied at "
            "the protein level. Verify the exact nt change in the Bioinformatic "
            "track.</span>"
        ).format(mn=spliceai_min, mx=spliceai)
    elif spliceai and spliceai > 0:
        if splice_pp3_active and splice_overrode_bp4:
            splice_section = (
                "<br><b>SpliceAI:</b> {sa:.2f} &mdash; "
                "<span style='color:#c00000;font-weight:bold'>"
                "splicing PP3 applies (supersedes missense BP4); "
                "potential splice disruption."
                "</span>"
            ).format(sa=spliceai)
        elif splice_pp3_active:
            splice_section = (
                "<br><b>SpliceAI:</b> {sa:.2f} &mdash; splicing PP3 applies"
            ).format(sa=spliceai)
        else:
            splice_section = "<br><b>SpliceAI:</b> {sa:.2f}".format(sa=spliceai)
    ps3_suppress_section = ""
    if ps3_suppressed:
        ps3_suppress_section = (
            "<br><b>Note:</b> {c} functional evidence not applied &mdash; "
            "splicing PP3 takes precedence per CSpec."
        ).format(c=ps3_suppressed)
    af_divergent_section = ""
    if af_divergent:
        af_divergent_section = (
            "<br><b>gnomAD AF (path-dependent):</b> "
            "<span style='color:#c00000'>the nt variants at this codon differ in "
            "allele frequency; verify the exact change in the Allele Frequencies "
            "track.</span>"
        )
    codon72 = ""
    if codon == 72:
        codon72 = (
            "<br><b>Note:</b> Codon 72 PS3/BS3 data is measured on the R72 "
            "haplotype (rs1042522); reference is P72. P72X variants are NOT "
            "in Table S3 (only R72X is). See description page for details."
        )
    ba1_section = ""
    if ba1:
        ba1_section = (
            "<br><b style='color:#0a0'>BA1 stand-alone Benign:</b> "
            "FAF &ge; 0.001 in gnomAD v4.1 forces class = Benign."
        )
    return (
        "{warn}"
        "<br><b>Variant:</b> p.{name} (NP_000537.3)"
        "<br><b>Provisional class:</b> {cls} ({pts} pts)"
        "<br><b>PM1:</b> {pm1}"
        "<br><b>PS3/BS3:</b> {ps3}"
        "<br><b>PP3/BP4:</b> {pp3}"
        "<br><b>AF (gnomAD v4.1):</b> {af}"
        "<br><b>BS2 (FLOSSIES):</b> {bs2}"
        "<br><b>Applied codes:</b> {applied}"
        "{splice}{ps3sup}{afdiv}{ba1_section}{codon72}"
        "<br><b>NOT included in this sum:</b> {cav}"
    ).format(warn=HEADER_WARNING,
             name=name, cls=cls, pts=pts,
             pm1=pm1 or "No contribution",
             ps3=ps3 or "No contribution",
             pp3=pp3 or "No contribution",
             af=af_lbl or "No contribution",
             bs2=bs2_lbl,
             applied=applied or "(none)",
             splice=splice_section,
             ps3sup=ps3_suppress_section,
             afdiv=af_divergent_section,
             ba1_section=ba1_section,
             codon72=codon72,
             cav=CAVEATS_STR)


def generate_bed(s3, s2, hotspot_occ, af_lookup, present_set, flossies_lookup, tx):
    lines = []
    chrom = tx['chrom']
    for (wt, codon, alt), s3_rec in sorted(s3.items(), key=lambda kv: (kv[0][1], kv[0][2])):
        pm1_lbl, pm1_pts = pm1_code_and_points(wt, codon, alt, hotspot_occ)
        ps3_lbl, ps3_pts = ps3_bs3_label_and_points(s3_rec['code'])
        s2_rec = s2.get((wt, codon, alt))
        pp3_lbl, pp3_pts = ('', 0)
        spliceai = 0.0
        spliceai_min = 0.0
        paths = []
        if s2_rec:
            pp3_lbl, pp3_pts = pp3_bp4_label_and_points(s2_rec['code'])
            spliceai = s2_rec.get('spliceai', 0.0)
            spliceai_min = s2_rec.get('spliceai_min', spliceai)
            paths = s2_rec.get('paths', [])

        # Splicing rule (SpliceAI >= 0.2 -> PP3 splicing, superseding a missense
        # BP4), applied only when UNAMBIGUOUS: every nt path yielding this protein
        # change is >= threshold. SpliceAI is nt-specific; when the paths straddle
        # the threshold this protein-level row cannot represent it, so we keep the
        # missense code and flag the row for per-nt verification (Bioinformatic).
        splice_all = bool(paths) and spliceai_min >= SPLICE_PP3_THRESHOLD
        splice_straddle = (spliceai >= SPLICE_PP3_THRESHOLD
                           and spliceai_min < SPLICE_PP3_THRESHOLD)
        splice_pp3_active = splice_all
        splice_overrode_bp4 = False
        if splice_pp3_active:
            if pp3_lbl in ('BP4', 'BP4_Moderate'):
                pp3_lbl = 'PP3 (splicing)'
                pp3_pts = 1
                splice_overrode_bp4 = True
            elif not pp3_lbl:
                pp3_lbl = 'PP3 (splicing)'
                pp3_pts = 1

        # CSpec caveat: do not apply a protein-level functional assay (PS3/BS3)
        # when the variant is classified as splicing via PP3.
        ps3_suppressed = ''
        if splice_pp3_active and ps3_lbl:
            ps3_suppressed = ps3_lbl
            ps3_lbl, ps3_pts = ('', 0)

        # Allele-frequency code (BA1 / BS1 / PM2_Supporting)
        af_code, af_pts, af_qty, af_divergent = af_code_and_points(
            wt, codon, alt, paths, af_lookup, present_set, tx)
        af_lbl = "{} ({})".format(af_code, af_qty) if af_code else ''
        ba1 = (af_code == 'BA1')

        # BS2 from FLOSSIES (tiered by carrier count per CSpec GN009)
        bs2_label, bs2_pts = bs2_evidence(wt, codon, paths, flossies_lookup, tx)
        bs2_applies = bs2_label is not None
        bs2_lbl = "{} ({} pts)".format(bs2_label, bs2_pts) if bs2_applies else "Not observed"

        total = pm1_pts + ps3_pts + pp3_pts + af_pts + bs2_pts

        if ba1:
            cls = 'Benign'
        else:
            cls = bucket(total)

        color = CLASS_COLOR[cls]

        # nt-divergence flags: this protein-level row summarizes >1 nt variant
        # whose SpliceAI or gnomAD AF disagree; verify the exact nt change in the
        # Bioinformatic / Allele Frequency tracks.
        nt_flags = []
        if splice_straddle:
            nt_flags.append('SpliceAI')
        if af_divergent:
            nt_flags.append('gnomAD AF')
        nt_verify = " + ".join(nt_flags) if nt_flags else "-"

        applied_codes = []
        if pm1_lbl: applied_codes.append("{} (+{})".format(pm1_lbl, pm1_pts))
        if ps3_lbl:
            sign = "+" if ps3_pts > 0 else ""
            applied_codes.append("{} ({}{})".format(ps3_lbl, sign, ps3_pts))
        if pp3_lbl:
            sign = "+" if pp3_pts > 0 else ""
            applied_codes.append("{} ({}{})".format(pp3_lbl, sign, pp3_pts))
        if af_code == 'BA1':
            applied_codes.append("BA1 (stand-alone B)")
        elif af_code:
            sign = "+" if af_pts > 0 else ""
            applied_codes.append("{} ({}{})".format(af_code, sign, af_pts))
        if bs2_applies:
            applied_codes.append("{} ({})".format(bs2_label, bs2_pts))
        applied = "; ".join(applied_codes)

        short = "{}{}{}".format(wt, codon, alt)
        mo = mouseover(short, cls, total,
                       pm1_lbl, ps3_lbl, pp3_lbl, af_lbl, bs2_lbl, applied,
                       spliceai, splice_pp3_active, splice_overrode_bp4,
                       codon, ba1, spliceai_min, splice_straddle, af_divergent,
                       ps3_suppressed)

        segs = lib.aa_codon_genomic(codon, tx)
        for g_start, g_end, _ex in segs:
            lines.append("\t".join([
                chrom, str(g_start), str(g_end),
                short, "0", ".",
                str(g_start), str(g_end),
                color,
                cls, str(total),
                applied,
                pm1_lbl or "-",
                ps3_lbl or "-",
                pp3_lbl or "-",
                af_code or "-",
                bs2_label if bs2_applies else "-",
                "{:.2f}".format(spliceai) if spliceai else "0.00",
                nt_verify,
                mo,
            ]))
    return lines


def build(db, outdir):
    print("=== {} ===".format(db))
    os.makedirs(outdir, exist_ok=True)
    s3 = load_s3(SRC_S3)
    s2 = load_s2(SRC_S2)
    hotspots = load_hotspot_occurrences(HOTSPOTS_JSON)
    af_lookup = load_af_lookup(db)
    present_set = load_gnomad_present(db)
    flossies_lookup = load_flossies_lookup(db)
    print("  S3 entries: {}   S2 entries: {}   "
          "cancerhotspots: {}   AF: {}   gnomAD present: {}   FLOSSIES BS2: {}".format(
              len(s3), len(s2), len(hotspots),
              len(af_lookup), len(present_set), len(flossies_lookup)))

    tx = lib.get_transcript_info(db)
    bed_lines = generate_bed(s3, s2, hotspots, af_lookup, present_set, flossies_lookup, tx)
    print("  {} BED rows".format(len(bed_lines)))

    as_file = os.path.join(outdir, "TP53ProvisionalClass.as")
    lib.write_autosql(as_file, AUTOSQL)
    bed = os.path.join(outdir, "TP53ProvisionalClass_{}.bed".format(db))
    with open(bed, 'w') as f:
        f.write("\n".join(bed_lines) + "\n")
    lib.run_sort_bed(bed)
    bb = os.path.join(outdir, "TP53ProvisionalClass{}.bb".format(db.capitalize()))
    lib.run_bedToBigBed(bed, as_file, bb, lib.chrom_sizes_path(db), "bed9+11")
    print("  wrote {}".format(bb))

    from collections import Counter
    cnt = Counter()
    af_cnt = Counter()
    bs2_cnt = Counter()
    with open(bed) as f:
        for line in f:
            flds = line.split("\t")
            cnt[flds[9]] += 1
            af_cnt[flds[15]] += 1
            if flds[16] in BS2_TIER_POINTS:
                bs2_cnt[flds[16]] += 1
    print("  Class distribution:")
    for k, n in cnt.most_common():
        print("    {}: {}".format(k, n))
    print("  AF code distribution: {}".format(dict(af_cnt)))
    print("  BS2 applied: {} rows {}".format(sum(bs2_cnt.values()), dict(bs2_cnt)))


def main():
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument('-o', '--output-dir', default=DEFAULT_OUTDIR)
    p.add_argument('--db', action='append', help='hg38 or hg19 (repeat). Default hg38.')
    args = p.parse_args()
    dbs = args.db if args.db else ['hg38']
    for db in dbs:
        build(db, args.output_dir)


if __name__ == "__main__":
    main()
